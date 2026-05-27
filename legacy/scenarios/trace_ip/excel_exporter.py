import json
import logging
import os

from scenarios.trace_ip.trace_utils import (
    LABEL_MAP,
    cat_display,
    extract_all_domains,
    extract_fofa_ports,
    has_domains,
    has_ports,
    is_china_ip,
    sort_key,
    trace_action,
    trace_priority,
)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logger = logging.getLogger('ip_info_manager.scenarios.trace_ip')

HEADERS = [
    'IP', '国家', 'ASN/组织', '分类', '分类说明', '建议溯源路径',
    '域名数', '反查域名列表', '端口数', '开放端口列表', '实时扫描端口数', '实时开放端口列表', '标签',
]

SHEET_CONFIG = {
    1: {'title': 'P1 核心溯源', 'desc': '有反查域名 + 国内IP'},
    2: {'title': 'P2 重点溯源', 'desc': '有反查域名（国外）或无域名但有端口（国内）'},
    3: {'title': 'P3 辅助溯源', 'desc': '无域名但有端口（国外）或仅国内IP'},
    4: {'title': 'P4 暂缓', 'desc': '无域名、无端口、国外IP'},
}


def _extract_domain_names(info):
    return [d['domain'] for d in extract_all_domains(info)]


def _extract_port_strings(info):
    ports = extract_fofa_ports(info)
    result = []
    for p in ports:
        port_str = str(p.get('port', ''))
        products = p.get('products', '')
        if products:
            port_str = f"{port_str}({products})"
        result.append(port_str)
    return result


def _cat_note(info):
    classify = info.get('trace_classify', {})
    matched_by = classify.get('matched_by', [])
    if matched_by and matched_by[0].get('note'):
        return matched_by[0]['note']
    return ''


def _trace_action_compat(info):
    return trace_action(info)


def _format_domain_with_verify(domain, domain_verify):
    if not domain_verify or not domain_verify.get('results'):
        return domain
    for r in domain_verify['results']:
        if r['domain'] == domain:
            status = r.get('status', '')
            if status == 'matched':
                return f'{domain} ✅'
            elif status == 'changed':
                ips = ', '.join(r.get('resolved_ips', []))
                return f'{domain} 🔄→{ips}'
            elif status == 'unresolved':
                return f'{domain} ❌'
            elif status == 'timeout':
                return f'{domain} ⏱️'
            elif status == 'error':
                return f'{domain} ⚠️'
            break
    return domain


def _extract_port_scan_ports(info):
    ps = info.get('port_scan', {})
    if not ps or 'error' in ps:
        return []
    return ps.get('open_ports', [])


def _build_row(ip, info):
    country = info.get('ipinfo_api', {}).get('country', '')
    org = info.get('ipinfo_api', {}).get('as_name', '')
    domains = _extract_domain_names(info)
    ports = _extract_port_strings(info)
    port_scan_ports = _extract_port_scan_ports(info)
    port_scan_strs = []
    for p in port_scan_ports:
        port_str = str(p.get('port', ''))
        service = p.get('service', '')
        product = p.get('product', '')
        if product:
            port_str = f"{port_str}({product})"
        elif service:
            port_str = f"{port_str}({service})"
        port_scan_strs.append(port_str)
    tags_data = info.get('tags', [])
    tags_str = ', '.join(tags_data) if isinstance(tags_data, list) else str(tags_data)
    domain_verify = info.get('domain_verify')
    if domain_verify and domain_verify.get('results'):
        formatted_domains = [_format_domain_with_verify(d, domain_verify) for d in domains]
    else:
        formatted_domains = domains
    return [
        ip,
        country,
        org,
        cat_display(info),
        _cat_note(info),
        _trace_action_compat(info),
        str(len(domains)),
        '\n'.join(formatted_domains),
        str(len(ports)),
        '\n'.join(ports),
        str(len(port_scan_ports)),
        '\n'.join(port_scan_strs),
        tags_str,
    ]


def generate_trace_excel(output_dir, prefix, exclude_info=None):
    if not OPENPYXL_AVAILABLE:
        logger.warning('openpyxl 未安装，跳过 Excel 导出。安装命令：pip install openpyxl')
        return False

    json_path = os.path.join(output_dir, f'{prefix}.json')
    if not os.path.exists(json_path):
        logger.warning('找不到数据文件 %s，跳过 Excel 导出', json_path)
        return False

    with open(json_path, 'r', encoding='utf-8') as f:
        ip_data = json.load(f)

    if exclude_info:
        exclude_set = exclude_info['exclude_ips']
        original_count = len(ip_data)
        ip_data = {ip: info for ip, info in ip_data.items() if ip not in exclude_set}
        logger.info("Excel 排除IP生效: 原始 %d 个, 排除 %d 个, 剩余 %d 个",
                    original_count, exclude_info['effective_count'], len(ip_data))

    deep_ips = []
    for ip in sorted(ip_data.keys()):
        if ip_data[ip].get('trace_classify', {}).get('need_deep_query'):
            deep_ips.append(ip)

    p_groups = {1: [], 2: [], 3: [], 4: []}
    for ip in deep_ips:
        lvl = trace_priority(ip_data[ip])
        p_groups[lvl].append(ip)

    for lvl in p_groups:
        p_groups[lvl].sort(key=lambda ip: sort_key(ip_data[ip]))

    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(
        start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_align = Alignment(
        horizontal='center', vertical='center', wrap_text=True)
    cell_align = Alignment(vertical='top', wrap_text=True)

    for lvl in [1, 2, 3, 4]:
        cfg = SHEET_CONFIG[lvl]
        ws = wb.create_sheet(title=cfg['title'])
        ws.append(HEADERS)

        for col_num in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        for ip in p_groups[lvl]:
            row = _build_row(ip, ip_data[ip])
            ws.append(row)
            row_num = ws.max_row
            for col_num in range(1, len(row) + 1):
                ws.cell(row=row_num, column=col_num).alignment = cell_align

        col_widths = {
            1: 18, 2: 10, 3: 25, 4: 22, 5: 20, 6: 30,
            7: 8, 8: 40, 9: 8, 10: 30, 11: 12, 12: 40, 13: 20,
        }
        for col_num, width in col_widths.items():
            ws.column_dimensions[chr(64 + col_num)].width = width

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = 'A2'

    output = os.path.join(output_dir, f'{prefix}.trace_report.xlsx')
    wb.save(output)
    logger.info('Excel 报告已生成：%s', output)
    return True
