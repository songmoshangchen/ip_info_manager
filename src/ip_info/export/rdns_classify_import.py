import json
import logging
import os
from collections import OrderedDict

logger = logging.getLogger(__name__)

VALID_MATCH_TYPES = {"suffix", "contains", "prefix", "exact", "regex"}

REQUIRED_COLS = [
    "is_sample",
    "hostname",
    "field",
    "category",
    "match_type",
    "match_value",
    "note",
    "skip",
    "new_label",
    "new_description",
    "new_need_deep_query",
]


def _cell_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _load_json_file(path: str) -> OrderedDict:
    if not os.path.exists(path):
        return OrderedDict()
    with open(path, "r", encoding="utf-8") as f:
        content = f.read().strip()
        if not content:
            return OrderedDict()
        return OrderedDict(json.loads(content))


def _save_json_file(path: str, data: OrderedDict):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _load_all_category_keys(rules_dir: str) -> set[str]:
    keys = set()
    for name in ("builtin_rules.json", "custom_rules.json"):
        path = os.path.join(rules_dir, name)
        if os.path.exists(path):
            keys.update(_load_json_file(path).keys())
    return keys


def validate_row(row: dict, existing_categories: set[str]) -> list[str]:
    errors = []
    is_sample = _cell_str(row.get("is_sample"))
    if is_sample:
        return errors
    skip = _cell_str(row.get("skip"))
    if skip:
        return errors

    category = _cell_str(row.get("category"))
    match_type = _cell_str(row.get("match_type"))
    match_value = _cell_str(row.get("match_value"))

    if not category:
        errors.append("category 不能为空")
    if not match_type:
        errors.append("match_type 不能为空")
    elif match_type not in VALID_MATCH_TYPES:
        errors.append(f"match_type 无效: {match_type}，有效值: {', '.join(sorted(VALID_MATCH_TYPES))}")
    if not match_value:
        errors.append("match_value 不能为空")

    if category and category not in existing_categories:
        new_label = _cell_str(row.get("new_label"))
        new_description = _cell_str(row.get("new_description"))
        new_need_deep_query = _cell_str(row.get("new_need_deep_query"))
        if not new_label:
            errors.append(f"新分类 '{category}' 必须填写 new_label")
        if not new_description:
            errors.append(f"新分类 '{category}' 必须填写 new_description")
        if new_need_deep_query not in ("是", "否"):
            errors.append(
                f"新分类 '{category}' 的 new_need_deep_query 必须为 '是' 或 '否'，当前值: '{new_need_deep_query}'"
            )

    return errors


def merge_rules(
    rows: list[dict],
    existing_categories: set[str],
    custom_rules: OrderedDict,
) -> tuple[OrderedDict, list[str]]:
    errors = []

    for row_idx, row in enumerate(rows, 1):
        is_sample = _cell_str(row.get("is_sample"))
        if is_sample:
            continue
        skip = _cell_str(row.get("skip"))
        if skip:
            continue

        row_errors = validate_row(row, existing_categories)
        if row_errors:
            errors.append(f"第 {row_idx} 行: {'; '.join(row_errors)}")
            continue

        category = _cell_str(row["category"])
        match_type = _cell_str(row["match_type"])
        match_value = _cell_str(row["match_value"])
        note = _cell_str(row.get("note"))
        field = _cell_str(row.get("field")) or "rdns_ptr.hostname"

        existing_patterns = set()
        if category in custom_rules:
            for p in custom_rules[category].get("patterns", []):
                existing_patterns.add(p.get("match", "").lower())

        if match_value.lower() in existing_patterns:
            continue

        new_pattern = {
            "field": field,
            "match": match_value,
            "type": match_type,
            "note": note,
        }

        if category in custom_rules:
            custom_rules[category]["patterns"].append(new_pattern)
        else:
            new_label = _cell_str(row.get("new_label"))
            new_description = _cell_str(row.get("new_description"))
            new_need_deep_query = _cell_str(row.get("new_need_deep_query")) == "是"

            custom_rules[category] = OrderedDict(
                [
                    ("label", new_label),
                    ("description", new_description),
                    ("need_deep_query", new_need_deep_query),
                    ("patterns", [new_pattern]),
                ]
            )
            existing_categories.add(category)

    return custom_rules, errors


def import_rdns_rules(
    excel_path: str,
    rules_dir: str,
    dry_run: bool = False,
) -> tuple[int, list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.error("openpyxl 未安装，无法读取 Excel")
        return 0, ["openpyxl 未安装"]

    if not os.path.exists(excel_path):
        return 0, [f"文件不存在: {excel_path}"]

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb["未分类RDNS"]

    headers = [cell.value for cell in ws[1]]
    col_map = {}
    for idx, header in enumerate(headers):
        if header and header in REQUIRED_COLS:
            col_map[header] = idx

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for col_name, col_idx in col_map.items():
            row_dict[col_name] = row[col_idx] if col_idx < len(row) else None
        rows.append(row_dict)

    wb.close()

    existing_categories = _load_all_category_keys(rules_dir)
    custom_path = os.path.join(rules_dir, "custom_rules.json")
    custom_rules = _load_json_file(custom_path)

    custom_rules, errors = merge_rules(rows, existing_categories, custom_rules)

    if errors:
        for err in errors:
            logger.error("验证错误: %s", err)

    if not dry_run and not errors:
        _save_json_file(custom_path, custom_rules)
        logger.info("已更新: %s", custom_path)
    elif dry_run:
        logger.info("[dry-run] 未写入文件，验证通过 %d 行", len(rows))

    added = sum(1 for r in rows if not _cell_str(r.get("is_sample")) and _cell_str(r.get("category")))
    return added, errors
