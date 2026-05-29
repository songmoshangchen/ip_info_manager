import json
import logging
import os

from ip_info.store.protocols import IPDataReader

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logger = logging.getLogger(__name__)

EXISTING_CATEGORIES = [
    "cloud_provider",
    "cdn",
    "crawler_scanner",
    "residential",
    "invalid_rdns",
    "excluded_domain",
]

CATEGORY_LABELS = {
    "cloud_provider": "云服务商",
    "cdn": "CDN/WAF",
    "crawler_scanner": "爬虫/扫描器",
    "residential": "家用宽带",
    "invalid_rdns": "无效RDNS",
    "excluded_domain": "排除域名",
}

SHEET1_HEADERS = [
    "is_sample",
    "hostname",
    "field",
    "category",
    "match_type",
    "match_value",
    "note",
    "skip",
    "new_label",
    "new_description",
    "new_need_deep_query",
]

SHEET1_WIDTHS = {
    "A": 10,
    "B": 35,
    "C": 20,
    "D": 20,
    "E": 12,
    "F": 25,
    "G": 25,
    "H": 8,
    "I": 15,
    "J": 25,
    "K": 20,
}

INSTRUCTION_ROWS = [
    [
        "说明",
        "RDNS反解域名",
        "匹配字段(固定)",
        "分类key",
        "匹配类型",
        "匹配值",
        "备注说明",
        "填跳过则不导入",
        "新分类时必填",
        "新分类时必填",
        "新分类时必填:是/否",
    ],
]

SAMPLE_ROWS = [
    [
        "样例",
        "ec2-1-2-3-4.amazonaws.com",
        "rdns_ptr.hostname",
        "cloud_provider",
        "suffix",
        ".amazonaws.com",
        "AWS Amazon 云服务",
        "",
        "",
        "",
        "",
    ],
    [
        "样例",
        "deuschef2",
        "rdns_ptr.hostname",
        "excluded_domain",
        "exact",
        "deuschef2",
        "单节无效域名",
        "",
        "",
        "",
        "",
    ],
    [
        "样例",
        "1.2.3.4",
        "rdns_ptr.hostname",
        "invalid_rdns",
        "regex",
        "^\\d{1,3}\\.\\d{1,3}\\.\\d{1,3}\\.\\d{1,3}$",
        "纯IP地址格式",
        "",
        "",
        "",
        "",
    ],
    [
        "样例",
        "dynamic-ip.isp.net",
        "rdns_ptr.hostname",
        "residential",
        "contains",
        ".dynamic",
        "动态 IP",
        "",
        "",
        "",
        "",
    ],
    [
        "样例",
        "crawl.baidu.com",
        "rdns_ptr.hostname",
        "crawler_scanner",
        "suffix",
        ".crawl.baidu.com",
        "百度搜索爬虫",
        "",
        "",
        "",
        "",
    ],
    [
        "样例",
        "cdn-node.cloudflare.com",
        "rdns_ptr.hostname",
        "cdn",
        "suffix",
        ".cloudflare.com",
        "Cloudflare CDN",
        "",
        "",
        "",
        "",
    ],
    [
        "样例",
        "my-custom-host.net",
        "rdns_ptr.hostname",
        "my_new_category",
        "suffix",
        ".my-custom-host.net",
        "自定义分类示例",
        "",
        "我的分类",
        "自定义分类描述",
        "否",
    ],
    [
        "样例",
        "same-owner-srv.example.com",
        "rdns_ptr.hostname",
        "cloud_provider",
        "suffix",
        ".example.com",
        "与上行同主域名",
        "跳过",
        "",
        "",
        "",
    ],
]

SHEET2_HEADERS = ["category", "label", "match_type", "match_value", "note", "example_hostname"]
SHEET2_WIDTHS = {"A": 18, "B": 12, "C": 12, "D": 25, "E": 25, "F": 35}


def _apply_header_style(ws, row_num: int, num_cols: int):
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_num in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align


def _apply_cell_align(ws, row_num: int, num_cols: int):
    cell_align = Alignment(vertical="top", wrap_text=True)
    for col_num in range(1, num_cols + 1):
        ws.cell(row=row_num, column=col_num).alignment = cell_align


def _set_column_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def _extract_unclassified_hostnames(reader: IPDataReader) -> list[str]:
    all_data = reader.list_all_ips_data()
    hostnames = set()
    for ip, info in all_data.items():
        classifier = info.get("classifier") or {}
        if classifier.get("category") != "other":
            continue
        rdns = info.get("rdns_ptr") or {}
        if not rdns.get("has_ptr"):
            continue
        hostname = rdns.get("hostname", "")
        if hostname:
            hostnames.add(hostname)
    return sorted(hostnames)


def _load_builtin_samples(rules_dir: str) -> list[dict]:
    builtin_path = os.path.join(rules_dir, "builtin_rules.json")
    if not os.path.exists(builtin_path):
        return []
    with open(builtin_path, "r", encoding="utf-8") as f:
        builtin = json.load(f)

    samples = []
    for cat_key in EXISTING_CATEGORIES:
        cat_def = builtin.get(cat_key)
        if not cat_def:
            continue
        patterns = cat_def.get("patterns", [])
        count = 0
        for pattern in patterns:
            if pattern.get("field") != "rdns_ptr.hostname":
                continue
            samples.append(
                {
                    "category": cat_key,
                    "label": cat_def.get("label", cat_key),
                    "match_type": pattern.get("type", "contains"),
                    "match_value": pattern.get("match", ""),
                    "note": pattern.get("note", ""),
                    "example_hostname": _generate_example_hostname(pattern),
                }
            )
            count += 1
            if count >= 3:
                break
    return samples


def _generate_example_hostname(pattern: dict) -> str:
    match_type = pattern.get("type", "contains")
    match_value = pattern.get("match", "")
    if match_type == "suffix":
        return f"host{match_value}"
    elif match_type == "contains":
        return f"prefix{match_value}suffix"
    elif match_type == "prefix":
        return f"{match_value}rest"
    elif match_type == "exact":
        return match_value
    elif match_type == "regex":
        return "1.2.3.4" if "\\d" in match_value else "example"
    return match_value


def _build_sheet1(ws, hostnames: list[str]):
    sample_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    sample_font = Font(color="806000")
    instruction_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    instruction_font = Font(color="375623", italic=True)

    ws.append(SHEET1_HEADERS)
    _apply_header_style(ws, 1, len(SHEET1_HEADERS))

    for row_data in INSTRUCTION_ROWS:
        ws.append(row_data)
        row_num = ws.max_row
        for col_num in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.fill = instruction_fill
            cell.font = instruction_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for row_data in SAMPLE_ROWS:
        ws.append(row_data)
        row_num = ws.max_row
        for col_num in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.fill = sample_fill
            cell.font = sample_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for hostname in hostnames:
        ws.append(["", hostname, "rdns_ptr.hostname", "", "", "", "", "", "", "", ""])
        _apply_cell_align(ws, ws.max_row, len(SHEET1_HEADERS))

    _set_column_widths(ws, SHEET1_WIDTHS)
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


def _build_sheet2(ws, samples: list[dict]):
    ws.append(SHEET2_HEADERS)
    _apply_header_style(ws, 1, len(SHEET2_HEADERS))

    current_cat = None
    cat_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    cat_font = Font(bold=True, size=11)

    for sample in samples:
        if sample["category"] != current_cat:
            current_cat = sample["category"]
            label = CATEGORY_LABELS.get(current_cat, current_cat)
            ws.append([current_cat, label, "", "", "", ""])
            row_num = ws.max_row
            for col_num in range(1, len(SHEET2_HEADERS) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.fill = cat_fill
                cell.font = cat_font
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        ws.append(
            [
                sample["category"],
                sample["label"],
                sample["match_type"],
                sample["match_value"],
                sample["note"],
                sample["example_hostname"],
            ]
        )
        _apply_cell_align(ws, ws.max_row, len(SHEET2_HEADERS))

    _set_column_widths(ws, SHEET2_WIDTHS)
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


def export_unclassified_rdns(
    reader: IPDataReader,
    output_dir: str,
    prefix: str,
    rules_dir: str,
) -> int:
    if not OPENPYXL_AVAILABLE:
        logger.warning("openpyxl 未安装，跳过未分类 RDNS 导出")
        return 0

    hostnames = _extract_unclassified_hostnames(reader)
    if not hostnames:
        logger.info("所有 RDNS 均已分类，无需导出")
        return 0

    samples = _load_builtin_samples(rules_dir)

    wb = Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet(title="未分类RDNS")
    _build_sheet1(ws1, hostnames)

    ws2 = wb.create_sheet(title="参考样例")
    _build_sheet2(ws2, samples)

    output_path = os.path.join(output_dir, f"{prefix}.unclassified_rdns.xlsx")
    wb.save(output_path)
    logger.info("未分类 RDNS Excel 已生成: %s (%d 个未分类 RDNS)", output_path, len(hostnames))
    return len(hostnames)
