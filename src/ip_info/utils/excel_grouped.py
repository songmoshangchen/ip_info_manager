import logging
import os
from dataclasses import dataclass
from typing import Callable

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logger = logging.getLogger(__name__)


@dataclass
class ColumnDef:
    header: str
    extract: Callable[[str, dict], str]
    width: int = 10


@dataclass
class SheetDef:
    title: str
    header_color: str = "4472C4"


@dataclass
class GroupLogic:
    group: Callable[[str, dict], int]
    sort_key: Callable[[str, dict], tuple]
    mark_fn: Callable[[str, dict], bool] | None = None


def generate_grouped_excel(
    ip_data: dict[str, dict],
    output_dir: str,
    prefix: str,
    columns: list[ColumnDef],
    sheets: list[SheetDef],
    logic: GroupLogic,
    mark_column: int = 1,
) -> bool:
    if not OPENPYXL_AVAILABLE:
        logger.warning("openpyxl 未安装，跳过 Excel 导出")
        return False

    groups: dict[int, list[str]] = {}
    for idx in range(len(sheets)):
        groups[idx] = []

    for ip in ip_data:
        info = ip_data[ip]
        idx = logic.group(ip, info)
        if idx < 0 or idx >= len(sheets):
            idx = len(sheets) - 1
        groups[idx].append(ip)

    for idx in groups:
        groups[idx].sort(key=lambda ip: logic.sort_key(ip, ip_data[ip]))

    wb = Workbook()
    wb.remove(wb.active)

    cell_align = Alignment(vertical="top", wrap_text=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_font = Font(bold=True, color="FFFFFF", size=11)

    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    green_font = Font(bold=True, color="006100")
    gray_cell_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    gray_cell_font = Font(color="808080")

    for idx, sheet_def in enumerate(sheets):
        ws = wb.create_sheet(title=sheet_def.title)

        header_fill = PatternFill(
            start_color=sheet_def.header_color,
            end_color=sheet_def.header_color,
            fill_type="solid",
        )
        headers = [col.header for col in columns]
        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        for ip in groups[idx]:
            info = ip_data[ip]
            row_values = [col.extract(ip, info) for col in columns]
            ws.append(row_values)
            row_num = ws.max_row
            for col_num in range(1, len(row_values) + 1):
                ws.cell(row=row_num, column=col_num).alignment = cell_align

            if logic.mark_fn is not None and 0 <= mark_column < len(columns):
                mark_cell = ws.cell(row=row_num, column=mark_column + 1)
                if logic.mark_fn(ip, info):
                    mark_cell.fill = green_fill
                    mark_cell.font = green_font
                else:
                    mark_cell.fill = gray_cell_fill
                    mark_cell.font = gray_cell_font

        for col_num, col_def in enumerate(columns, 1):
            ws.column_dimensions[chr(64 + col_num)].width = col_def.width

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

    output = os.path.join(output_dir, f"{prefix}.xlsx")
    wb.save(output)
    logger.info("Excel 已生成：%s", output)
    return True
