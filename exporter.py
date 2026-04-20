import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


class IPExcelExporter:
    def __init__(self, reader):
        self.reader = reader
    
    def flatten_json(self, data, prefix=''):
        items = []
        for k, v in data.items():
            new_key = f"{prefix}.{k}" if prefix else k
            if isinstance(v, dict):
                items.extend(self.flatten_json(v, new_key).items())
            elif isinstance(v, list):
                items.append((new_key, ', '.join(map(str, v))))
            else:
                items.append((new_key, v))
        return dict(items)
    
    def collect_fields_from_data(self, ips_data):
        all_fields = set()
        for ip, data in ips_data.items():
            flat = self.flatten_json(data)
            flat.pop('ip', None)
            all_fields.update(flat.keys())
        return sorted(all_fields)
    
    def confirm_fields(self, fields, ip_count):
        print(f"\n即将导出 {ip_count} 个 IP，共 {len(fields)} 个字段:")
        print("=" * 60)
        
        current_channel = None
        for field in fields:
            parts = field.split('.', 1)
            channel = parts[0] if parts else field
            
            if channel != current_channel:
                current_channel = channel
                print(f"\n【{channel}】")
            
            field_path = parts[1] if len(parts) > 1 else ''
            if field_path:
                print(f"  - {field_path}")
            else:
                print(f"  - {field}")
        
        print("\n" + "=" * 60)
        while True:
            response = input("是否继续导出？[y/n]: ").strip().lower()
            if response in ['y', 'yes', '是']:
                return True
            elif response in ['n', 'no', '否']:
                return False
            print("请输入 y 或 n")
    
    def export_to_excel(self, ips, include_channel=None, exclude_channel=None, output_file=None):
        if not ips:
            print("没有数据可导出")
            return
        
        ips_data = {}
        for ip in ips:
            ip_data = self.reader.get_ip_data(ip)
            if ip_data:
                ips_data[ip] = ip_data
        
        if not ips_data:
            print("没有有效数据可导出")
            return
        
        filtered_ips_data = {}
        for ip, data in ips_data.items():
            if include_channel:
                filtered_data = {'ip': data.get('ip', ip)}
                for ch in include_channel:
                    if ch in data:
                        filtered_data[ch] = data[ch]
                filtered_ips_data[ip] = filtered_data
            elif exclude_channel:
                filtered_data = {k: v for k, v in data.items() if k not in exclude_channel}
                filtered_ips_data[ip] = filtered_data
            else:
                filtered_ips_data[ip] = data
        
        ips_data = filtered_ips_data
        
        all_fields = self.collect_fields_from_data(ips_data)
        
        if not self.confirm_fields(all_fields, len(ips_data)):
            print("已取消导出")
            return
        
        wb = Workbook()
        ws = wb.active
        ws.title = "IP Data"
        
        headers = ['IP'] + all_fields
        ws.append(headers)
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for ip in sorted(ips_data.keys()):
            ip_data = ips_data[ip]
            flat_data = self.flatten_json(ip_data)
            
            row_data = [ip]
            for field in all_fields:
                row_data.append(flat_data.get(field, ''))
            
            ws.append(row_data)
        
        for col_num in range(1, len(headers) + 1):
            max_length = 0
            column = ws.cell(row=1, column=col_num)
            
            for row_num in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_num, column=col_num)
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[chr(64 + col_num)].width = adjusted_width
        
        output_path = output_file or 'ip_data_export.xlsx'
        wb.save(output_path)
        print(f"\n导出成功！文件已保存至: {output_path}")
