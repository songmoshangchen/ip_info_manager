import os

import pytest

from ip_info.utils.excel_grouped import (
    ColumnDef,
    GroupLogic,
    SheetDef,
    generate_grouped_excel,
)


@pytest.fixture
def output_dir(tmp_path):
    return str(tmp_path)


def _read_xlsx(xlsx_path):
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path)
    sheets = {}
    for ws in wb.worksheets:
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        sheets[ws.title] = rows
    wb.close()
    return sheets


def _read_xlsx_styles(xlsx_path):
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path)
    result = {}
    for ws in wb.worksheets:
        result[ws.title] = {
            "header_fills": [ws.cell(1, c).fill.start_color.rgb for c in range(1, ws.max_column + 1)],
            "cell_values": [
                [ws.cell(r, c).value for c in range(1, ws.max_column + 1)] for r in range(2, ws.max_row + 1)
            ],
            "mark_fills": [ws.cell(r, 2).fill.start_color.rgb for r in range(2, ws.max_row + 1)]
            if ws.max_row > 1
            else [],
            "mark_fonts": [ws.cell(r, 2).font.color.rgb for r in range(2, ws.max_row + 1)] if ws.max_row > 1 else [],
        }
    wb.close()
    return result


class TestColumnDef:
    def test_extract_callable(self):
        col = ColumnDef(header="Name", extract=lambda ip, _: ip, width=15)
        assert col.extract("1.1.1.1", {}) == "1.1.1.1"
        assert col.header == "Name"
        assert col.width == 15


class TestBasicGeneration:
    def test_creates_xlsx(self, output_dir):
        columns = [ColumnDef("IP", lambda ip, _: ip)]
        sheets = [SheetDef("Group A"), SheetDef("Group B")]
        logic = GroupLogic(
            group=lambda ip, _: 0 if ip < "5" else 1,
            sort_key=lambda ip, _: (ip,),
        )
        ip_data = {"1.1.1.1": {}, "9.9.9.9": {}}

        result = generate_grouped_excel(ip_data, output_dir, "test", columns, sheets, logic)

        assert result is True
        assert os.path.exists(os.path.join(output_dir, "test.xlsx"))

    def test_sheet_titles(self, output_dir):
        columns = [ColumnDef("IP", lambda ip, _: ip)]
        sheets = [SheetDef("Sheet1"), SheetDef("Sheet2")]
        logic = GroupLogic(
            group=lambda ip, _: 0,
            sort_key=lambda ip, _: (ip,),
        )

        generate_grouped_excel({"1.1.1.1": {}}, output_dir, "test", columns, sheets, logic)
        data = _read_xlsx(os.path.join(output_dir, "test.xlsx"))

        assert list(data.keys()) == ["Sheet1", "Sheet2"]

    def test_headers_match_columns(self, output_dir):
        columns = [
            ColumnDef("IP", lambda ip, _: ip),
            ColumnDef("Country", lambda _, info: info.get("country", "")),
        ]
        sheets = [SheetDef("All")]
        logic = GroupLogic(group=lambda ip, _: 0, sort_key=lambda ip, _: (ip,))

        generate_grouped_excel({"1.1.1.1": {"country": "CN"}}, output_dir, "test", columns, sheets, logic)
        data = _read_xlsx(os.path.join(output_dir, "test.xlsx"))

        assert data["All"][0] == ["IP", "Country"]


class TestGrouping:
    def test_ips_grouped_by_logic(self, output_dir):
        columns = [ColumnDef("IP", lambda ip, _: ip)]
        sheets = [SheetDef("Low"), SheetDef("High")]
        logic = GroupLogic(
            group=lambda ip, _: 1 if ip.startswith("9") else 0,
            sort_key=lambda ip, _: (ip,),
        )
        ip_data = {"1.1.1.1": {}, "2.2.2.2": {}, "9.9.9.9": {}}

        generate_grouped_excel(ip_data, output_dir, "test", columns, sheets, logic)
        data = _read_xlsx(os.path.join(output_dir, "test.xlsx"))

        assert len(data["Low"]) == 3
        assert len(data["High"]) == 2
        assert data["Low"][1][0] == "1.1.1.1"
        assert data["Low"][2][0] == "2.2.2.2"
        assert data["High"][1][0] == "9.9.9.9"

    def test_sorting(self, output_dir):
        columns = [ColumnDef("IP", lambda ip, _: ip)]
        sheets = [SheetDef("All")]
        logic = GroupLogic(
            group=lambda ip, _: 0,
            sort_key=lambda ip, info: (-info.get("score", 0),),
        )
        ip_data = {"a": {"score": 1}, "b": {"score": 3}, "c": {"score": 2}}

        generate_grouped_excel(ip_data, output_dir, "test", columns, sheets, logic)
        data = _read_xlsx(os.path.join(output_dir, "test.xlsx"))

        assert data["All"][1][0] == "b"
        assert data["All"][2][0] == "c"
        assert data["All"][3][0] == "a"


class TestMarkColumn:
    def test_mark_true_green(self, output_dir):
        columns = [
            ColumnDef("IP", lambda ip, _: ip),
            ColumnDef("Flag", lambda ip, info: "是" if info.get("ok") else "否"),
        ]
        sheets = [SheetDef("All")]
        logic = GroupLogic(
            group=lambda ip, _: 0,
            sort_key=lambda ip, _: (ip,),
            mark_fn=lambda ip, info: info.get("ok", False),
        )

        generate_grouped_excel({"1.1.1.1": {"ok": True}}, output_dir, "test", columns, sheets, logic)
        styles = _read_xlsx_styles(os.path.join(output_dir, "test.xlsx"))

        assert styles["All"]["mark_fills"] == ["00E2EFDA"]
        assert styles["All"]["mark_fonts"] == ["00006100"]

    def test_mark_false_gray(self, output_dir):
        columns = [
            ColumnDef("IP", lambda ip, _: ip),
            ColumnDef("Flag", lambda ip, info: "是" if info.get("ok") else "否"),
        ]
        sheets = [SheetDef("All")]
        logic = GroupLogic(
            group=lambda ip, _: 0,
            sort_key=lambda ip, _: (ip,),
            mark_fn=lambda ip, info: info.get("ok", False),
        )

        generate_grouped_excel({"1.1.1.1": {"ok": False}}, output_dir, "test", columns, sheets, logic)
        styles = _read_xlsx_styles(os.path.join(output_dir, "test.xlsx"))

        assert styles["All"]["mark_fills"] == ["00F2F2F2"]
        assert styles["All"]["mark_fonts"] == ["00808080"]


class TestSheetStyling:
    def test_custom_header_color(self, output_dir):
        columns = [ColumnDef("IP", lambda ip, _: ip)]
        sheets = [SheetDef("Normal"), SheetDef("Gray", header_color="A6A6A6")]
        logic = GroupLogic(group=lambda ip, _: 0, sort_key=lambda ip, _: (ip,))

        generate_grouped_excel({}, output_dir, "test", columns, sheets, logic)
        styles = _read_xlsx_styles(os.path.join(output_dir, "test.xlsx"))

        assert styles["Normal"]["header_fills"][0] == "004472C4"
        assert styles["Gray"]["header_fills"][0] == "00A6A6A6"


class TestColumnValues:
    def test_extract_uses_ip_and_info(self, output_dir):
        columns = [
            ColumnDef("IP", lambda ip, _: ip),
            ColumnDef("Country", lambda _, info: info.get("country", "")),
            ColumnDef("Score", lambda _, info: str(info.get("score", 0))),
        ]
        sheets = [SheetDef("All")]
        logic = GroupLogic(group=lambda ip, _: 0, sort_key=lambda ip, _: (ip,))

        ip_data = {"1.1.1.1": {"country": "CN", "score": 42}}
        generate_grouped_excel(ip_data, output_dir, "test", columns, sheets, logic)
        data = _read_xlsx(os.path.join(output_dir, "test.xlsx"))

        assert data["All"][1] == ["1.1.1.1", "CN", "42"]


class TestEdgeCases:
    def test_empty_ip_data(self, output_dir):
        columns = [ColumnDef("IP", lambda ip, _: ip)]
        sheets = [SheetDef("All")]
        logic = GroupLogic(group=lambda ip, _: 0, sort_key=lambda ip, _: (ip,))

        result = generate_grouped_excel({}, output_dir, "test", columns, sheets, logic)
        assert result is True
        data = _read_xlsx(os.path.join(output_dir, "test.xlsx"))
        assert len(data["All"]) == 1

    def test_group_out_of_range_goes_to_last(self, output_dir):
        columns = [ColumnDef("IP", lambda ip, _: ip)]
        sheets = [SheetDef("A"), SheetDef("B")]
        logic = GroupLogic(
            group=lambda ip, _: 99,
            sort_key=lambda ip, _: (ip,),
        )

        generate_grouped_excel({"1.1.1.1": {}}, output_dir, "test", columns, sheets, logic)
        data = _read_xlsx(os.path.join(output_dir, "test.xlsx"))

        assert len(data["A"]) == 1
        assert len(data["B"]) == 2


class TestFreezeAndFilter:
    def test_freeze_and_autofilter(self, output_dir):
        columns = [ColumnDef("IP", lambda ip, _: ip)]
        sheets = [SheetDef("All")]
        logic = GroupLogic(group=lambda ip, _: 0, sort_key=lambda ip, _: (ip,))

        generate_grouped_excel({"1.1.1.1": {}}, output_dir, "test", columns, sheets, logic)

        from openpyxl import load_workbook

        wb = load_workbook(os.path.join(output_dir, "test.xlsx"))
        ws = wb["All"]
        assert ws.freeze_panes == "A2"
        assert ws.auto_filter.ref is not None
        wb.close()
