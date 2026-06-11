import json
import os

import pytest

from ip_info.export.trace_judge_excel import (
    ChannelMapping,
    generate_trace_judge_excel,
    generate_trace_only_excel,
)


@pytest.fixture
def output_dir(tmp_path):
    return str(tmp_path)


def _make_ip_info(
    ip,
    country="CN",
    country_code="CN",
    as_name="TestOrg",
    category="cloud_provider",
    need_deep_query=True,
    has_domains=True,
    has_fofa_ports=False,
    has_port_scan=False,
    has_rdns=False,
    tags=None,
):
    info = {
        "ipinfo_api": {
            "country": country,
            "country_code": country_code,
            "as_name": as_name,
        },
        "classifier": {
            "category": category,
            "need_deep_query": need_deep_query,
            "matched_by": [],
        },
    }
    if has_domains:
        info["chinaz"] = {
            "success": True,
            "domains": [{"domain": f"example-{ip}.com"}],
        }
    if has_fofa_ports:
        info["fofa_host"] = {"ports": [{"port": 80, "protocol": "http", "products": []}]}
    if has_port_scan:
        info["port_scan"] = {"open_ports": [{"port": 443, "service": "https"}]}
    if has_rdns:
        info["rdns_ptr"] = {"has_ptr": True, "hostname": f"host-{ip}"}
    if tags:
        info["tagger"] = {"tags": tags}
    return ip, info


def _write_json(output_dir, prefix, ip_data):
    json_path = os.path.join(output_dir, f"{prefix}.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(ip_data, f)
    return json_path


def _read_xlsx(xlsx_path):
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path)
    sheets = {}
    for ws in wb.worksheets:
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        sheets[ws.title] = rows
    wb.close()
    return sheets


# ===== 模式1: 全量输出（含需要溯源列）=====


class TestFullOutput:
    def test_creates_five_sheets(self, output_dir):
        ip1, info1 = _make_ip_info("1.1.1.1", has_domains=True, has_fofa_ports=True)
        ip2, info2 = _make_ip_info("2.2.2.2", category="crawler_scanner", need_deep_query=False, has_domains=False)
        _write_json(output_dir, "test", {ip1: info1, ip2: info2})

        result = generate_trace_judge_excel(output_dir, "test")
        assert result is True
        xlsx_path = os.path.join(output_dir, "test.trace_judge.xlsx")
        assert os.path.exists(xlsx_path)
        sheets = _read_xlsx(xlsx_path)
        assert list(sheets.keys()) == ["P1 核心溯源", "P2 重点溯源", "P3 辅助溯源", "P4 暂缓", "P5 不需溯源"]

    def test_crawler_scanner_in_p5(self, output_dir):
        ip, info = _make_ip_info("1.1.1.1", category="crawler_scanner", need_deep_query=False, has_domains=False)
        _write_json(output_dir, "test", {ip: info})

        generate_trace_judge_excel(output_dir, "test")
        sheets = _read_xlsx(os.path.join(output_dir, "test.trace_judge.xlsx"))
        assert len(sheets["P5 不需溯源"]) == 2

    def test_exclude_ip_stays_in_priority_marked_no(self, output_dir):
        ip, info = _make_ip_info("1.1.1.1", has_domains=True, has_fofa_ports=True)
        _write_json(output_dir, "test", {ip: info})

        generate_trace_judge_excel(output_dir, "test", exclude_ips={"1.1.1.1"})
        sheets = _read_xlsx(os.path.join(output_dir, "test.trace_judge.xlsx"))
        assert sheets["P1 核心溯源"][1][0] == "1.1.1.1"
        assert sheets["P1 核心溯源"][1][1] == "否"

    def test_normal_ip_marked_yes(self, output_dir):
        ip, info = _make_ip_info("1.1.1.1", has_domains=True)
        _write_json(output_dir, "test", {ip: info})

        generate_trace_judge_excel(output_dir, "test")
        sheets = _read_xlsx(os.path.join(output_dir, "test.trace_judge.xlsx"))
        assert sheets["P1 核心溯源"][1][1] == "是"

    def test_returns_false_when_no_json(self, output_dir):
        result = generate_trace_judge_excel(output_dir, "missing")
        assert result is False


# ===== 模式2: 仅需要溯源的IP =====


class TestTraceOnlyOutput:
    def test_creates_four_sheets_no_p5(self, output_dir):
        ip1, info1 = _make_ip_info("1.1.1.1", has_domains=True, has_fofa_ports=True)
        ip2, info2 = _make_ip_info("2.2.2.2", category="crawler_scanner", need_deep_query=False, has_domains=False)
        _write_json(output_dir, "test", {ip1: info1, ip2: info2})

        result = generate_trace_only_excel(output_dir, "test")
        assert result is True
        xlsx_path = os.path.join(output_dir, "test.trace_only.xlsx")
        assert os.path.exists(xlsx_path)
        sheets = _read_xlsx(xlsx_path)
        assert list(sheets.keys()) == ["P1 核心溯源", "P2 重点溯源", "P3 辅助溯源", "P4 暂缓"]

    def test_no_needs_trace_column(self, output_dir):
        ip, info = _make_ip_info("1.1.1.1", has_domains=True)
        _write_json(output_dir, "test", {ip: info})

        generate_trace_only_excel(output_dir, "test")
        sheets = _read_xlsx(os.path.join(output_dir, "test.trace_only.xlsx"))
        headers = sheets["P1 核心溯源"][0]
        assert "需要溯源" not in headers

    def test_excludes_crawler_scanner(self, output_dir):
        ip1, info1 = _make_ip_info("1.1.1.1", has_domains=True)
        ip2, info2 = _make_ip_info("2.2.2.2", category="crawler_scanner", need_deep_query=False, has_domains=False)
        _write_json(output_dir, "test", {ip1: info1, ip2: info2})

        generate_trace_only_excel(output_dir, "test")
        sheets = _read_xlsx(os.path.join(output_dir, "test.trace_only.xlsx"))
        total_ips = sum(len(sheets[t]) - 1 for t in sheets)
        assert total_ips == 1

    def test_excludes_exclude_ips(self, output_dir):
        ip, info = _make_ip_info("1.1.1.1", has_domains=True, has_fofa_ports=True)
        _write_json(output_dir, "test", {ip: info})

        generate_trace_only_excel(output_dir, "test", exclude_ips={"1.1.1.1"})
        sheets = _read_xlsx(os.path.join(output_dir, "test.trace_only.xlsx"))
        total_ips = sum(len(sheets[t]) - 1 for t in sheets)
        assert total_ips == 0

    def test_returns_false_when_no_json(self, output_dir):
        result = generate_trace_only_excel(output_dir, "missing")
        assert result is False


# ===== ChannelMapping =====


class TestChannelMapping:
    def test_default_values(self):
        ch = ChannelMapping()
        assert ch.ipinfo == "ipinfo_api"
        assert ch.classifier == "classifier"
        assert ch.domain_sources == ["aizhan", "chinaz"]
        assert ch.fofa_ports == "fofa_host"
        assert ch.port_scan == "port_scan"
        assert ch.tagger == "tagger"
        assert ch.rdns == "rdns_ptr"

    def test_custom_values(self):
        ch = ChannelMapping(ipinfo="whois", domain_sources=["sublist3r"])
        assert ch.ipinfo == "whois"
        assert ch.domain_sources == ["sublist3r"]


# ===== 优先级分组 =====


class TestPriorityGrouping:
    # ===== P1: malicious / 国内+域名 / 国内+端口 =====

    def test_malicious_is_p1(self, output_dir):
        ip, info = _make_ip_info("1.1.1.1", category="malicious", has_domains=False, has_fofa_ports=False)
        _write_json(output_dir, "test", {ip: info})
        generate_trace_judge_excel(output_dir, "test")
        sheets = _read_xlsx(os.path.join(output_dir, "test.trace_judge.xlsx"))
        assert len(sheets["P1 核心溯源"]) == 2

    def test_cn_ip_with_domains_is_p1(self, output_dir):
        ip, info = _make_ip_info("1.1.1.1", has_domains=True)
        _write_json(output_dir, "test", {ip: info})
        generate_trace_judge_excel(output_dir, "test")
        sheets = _read_xlsx(os.path.join(output_dir, "test.trace_judge.xlsx"))
        assert len(sheets["P1 核心溯源"]) == 2

    def test_cn_ip_with_ports_is_p1(self, output_dir):
        ip, info = _make_ip_info("1.1.1.1", has_domains=False, has_fofa_ports=True)
        _write_json(output_dir, "test", {ip: info})
        generate_trace_judge_excel(output_dir, "test")
        sheets = _read_xlsx(os.path.join(output_dir, "test.trace_judge.xlsx"))
        assert len(sheets["P1 核心溯源"]) == 2

    def test_cn_ip_with_port_scan_only_is_p1(self, output_dir):
        ip, info = _make_ip_info("1.1.1.1", has_domains=False, has_fofa_ports=False, has_port_scan=True)
        _write_json(output_dir, "test", {ip: info})
        generate_trace_judge_excel(output_dir, "test")
        sheets = _read_xlsx(os.path.join(output_dir, "test.trace_judge.xlsx"))
        assert len(sheets["P1 核心溯源"]) == 2

    # ===== P2: 国外+域名+服务器 / 国外+端口+服务器 / 国内+家宽 =====

    def test_foreign_ip_with_domains_cloud_is_p2(self, output_dir):
        ip, info = _make_ip_info(
            "2.2.2.2", country="US", country_code="US", has_domains=True, category="cloud_provider"
        )
        _write_json(output_dir, "test", {ip: info})
        generate_trace_judge_excel(output_dir, "test")
        sheets = _read_xlsx(os.path.join(output_dir, "test.trace_judge.xlsx"))
        assert len(sheets["P2 重点溯源"]) == 2

    def test_foreign_ip_with_ports_cloud_is_p2(self, output_dir):
        ip, info = _make_ip_info(
            "4.4.4.4",
            country="US",
            country_code="US",
            has_domains=False,
            has_fofa_ports=True,
            category="cloud_provider",
        )
        _write_json(output_dir, "test", {ip: info})
        generate_trace_judge_excel(output_dir, "test")
        sheets = _read_xlsx(os.path.join(output_dir, "test.trace_judge.xlsx"))
        assert len(sheets["P2 重点溯源"]) == 2

    def test_cn_residential_is_p2(self, output_dir):
        ip, info = _make_ip_info("3.3.3.3", category="residential", has_domains=False, has_fofa_ports=False)
        _write_json(output_dir, "test", {ip: info})
        generate_trace_judge_excel(output_dir, "test")
        sheets = _read_xlsx(os.path.join(output_dir, "test.trace_judge.xlsx"))
        assert len(sheets["P2 重点溯源"]) == 2

    # ===== P3: 国外+服务器(裸) / 国内+other(裸) =====

    def test_foreign_cloud_bare_is_p3(self, output_dir):
        ip, info = _make_ip_info(
            "5.5.5.5", country="US", country_code="US", has_domains=False, category="cloud_provider"
        )
        _write_json(output_dir, "test", {ip: info})
        generate_trace_judge_excel(output_dir, "test")
        sheets = _read_xlsx(os.path.join(output_dir, "test.trace_judge.xlsx"))
        assert len(sheets["P3 辅助溯源"]) == 2

    def test_cn_other_bare_is_p3(self, output_dir):
        ip, info = _make_ip_info("6.6.6.6", category="other", has_domains=False, has_fofa_ports=False)
        _write_json(output_dir, "test", {ip: info})
        generate_trace_judge_excel(output_dir, "test")
        sheets = _read_xlsx(os.path.join(output_dir, "test.trace_judge.xlsx"))
        assert len(sheets["P3 辅助溯源"]) == 2

    # ===== P4: 其余（国外+家宽/other） =====

    def test_foreign_residential_is_p4(self, output_dir):
        ip, info = _make_ip_info("7.7.7.7", country="US", country_code="US", category="residential", has_domains=False)
        _write_json(output_dir, "test", {ip: info})
        generate_trace_judge_excel(output_dir, "test")
        sheets = _read_xlsx(os.path.join(output_dir, "test.trace_judge.xlsx"))
        assert len(sheets["P4 暂缓"]) == 2

    def test_foreign_other_is_p4(self, output_dir):
        ip, info = _make_ip_info("8.8.8.8", country="US", country_code="US", category="other", has_domains=False)
        _write_json(output_dir, "test", {ip: info})
        generate_trace_judge_excel(output_dir, "test")
        sheets = _read_xlsx(os.path.join(output_dir, "test.trace_judge.xlsx"))
        assert len(sheets["P4 暂缓"]) == 2

    # ===== P5: 仅 crawler_scanner / cdn =====

    def test_cdn_is_p5(self, output_dir):
        ip, info = _make_ip_info("9.9.9.9", category="cdn", has_domains=True)
        _write_json(output_dir, "test", {ip: info})
        generate_trace_judge_excel(output_dir, "test")
        sheets = _read_xlsx(os.path.join(output_dir, "test.trace_judge.xlsx"))
        assert len(sheets["P5 不需溯源"]) == 2

    def test_need_deep_query_false_not_p5_if_not_noise(self, output_dir):
        """residential + need_deep_query=False 不再被排到 P5。"""
        ip, info = _make_ip_info("10.10.10.10", category="residential", need_deep_query=False, has_domains=False)
        _write_json(output_dir, "test", {ip: info})
        generate_trace_judge_excel(output_dir, "test")
        sheets = _read_xlsx(os.path.join(output_dir, "test.trace_judge.xlsx"))
        assert len(sheets["P5 不需溯源"]) == 1  # only header
