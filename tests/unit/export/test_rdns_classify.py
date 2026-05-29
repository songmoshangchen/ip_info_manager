import json
import os
from collections import OrderedDict

import pytest

from ip_info.export.rdns_classify_excel import export_unclassified_rdns
from ip_info.export.rdns_classify_import import merge_rules, validate_row
from ip_info.store.in_memory import InMemoryIPReader


@pytest.fixture
def output_dir(tmp_path):
    return str(tmp_path)


@pytest.fixture
def rules_dir(tmp_path):
    builtin = {
        "cloud_provider": {
            "label": "云服务商",
            "description": "公有云/私有云主机",
            "need_deep_query": True,
            "patterns": [
                {
                    "field": "rdns_ptr.hostname",
                    "match": ".amazonaws.com",
                    "type": "suffix",
                    "note": "AWS",
                },
                {
                    "field": "rdns_ptr.hostname",
                    "match": ".aliyun.com",
                    "type": "suffix",
                    "note": "阿里云",
                },
            ],
        },
        "cdn": {
            "label": "CDN/WAF",
            "description": "CDN或WAF节点",
            "need_deep_query": False,
            "patterns": [
                {
                    "field": "rdns_ptr.hostname",
                    "match": ".cloudflare.com",
                    "type": "suffix",
                    "note": "Cloudflare",
                },
            ],
        },
        "invalid_rdns": {
            "label": "无效RDNS",
            "description": "RDNS反解为纯IP地址格式",
            "need_deep_query": False,
            "patterns": [
                {
                    "field": "rdns_ptr.hostname",
                    "match": "^\\d{1,3}\\.\\d{1,3}\\.\\d{1,3}\\.\\d{1,3}$",
                    "type": "regex",
                    "note": "纯IP",
                },
            ],
        },
    }
    rules_path = os.path.join(str(tmp_path), "builtin_rules.json")
    with open(rules_path, "w", encoding="utf-8") as f:
        json.dump(builtin, f, ensure_ascii=False)

    custom_path = os.path.join(str(tmp_path), "custom_rules.json")
    with open(custom_path, "w", encoding="utf-8") as f:
        json.dump({}, f)

    return str(tmp_path)


def _make_reader(ip_data: dict) -> InMemoryIPReader:
    store = {}
    for ip, channels in ip_data.items():
        entry = {"ip": ip}
        entry.update(channels)
        store[ip] = entry
    return InMemoryIPReader(data=store)


def _read_xlsx(xlsx_path):
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path)
    sheets = {}
    for ws in wb.worksheets:
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        sheets[ws.title] = rows
    wb.close()
    return sheets


def _find_col(headers, name):
    for idx, h in enumerate(headers):
        if h == name:
            return idx
    return -1


# ===== Export 行为 =====


class TestExportUnclassifiedRdns:
    def test_generates_excel_with_correct_count(self, output_dir, rules_dir):
        data = {
            "1.1.1.1": {
                "classifier": {"category": "other"},
                "rdns_ptr": {"has_ptr": True, "hostname": "unknown.example.com"},
            },
            "2.2.2.2": {
                "classifier": {"category": "other"},
                "rdns_ptr": {"has_ptr": True, "hostname": "another.example.com"},
            },
        }
        reader = _make_reader(data)

        count = export_unclassified_rdns(reader, output_dir, "test", rules_dir)
        assert count == 2

        xlsx_path = os.path.join(output_dir, "test.unclassified_rdns.xlsx")
        assert os.path.exists(xlsx_path)

    def test_excel_has_two_sheets(self, output_dir, rules_dir):
        data = {
            "1.1.1.1": {
                "classifier": {"category": "other"},
                "rdns_ptr": {"has_ptr": True, "hostname": "unknown.example.com"},
            },
        }
        reader = _make_reader(data)
        export_unclassified_rdns(reader, output_dir, "test", rules_dir)

        sheets = _read_xlsx(os.path.join(output_dir, "test.unclassified_rdns.xlsx"))
        assert list(sheets.keys()) == ["未分类RDNS", "参考样例"]

    def test_sheet1_contains_instruction_sample_and_data_rows(self, output_dir, rules_dir):
        data = {
            "1.1.1.1": {
                "classifier": {"category": "other"},
                "rdns_ptr": {"has_ptr": True, "hostname": "unknown.example.com"},
            },
        }
        reader = _make_reader(data)
        export_unclassified_rdns(reader, output_dir, "test", rules_dir)

        sheets = _read_xlsx(os.path.join(output_dir, "test.unclassified_rdns.xlsx"))
        rows = sheets["未分类RDNS"]
        headers = rows[0]

        is_sample_col = _find_col(headers, "is_sample")
        hostname_col = _find_col(headers, "hostname")
        skip_col = _find_col(headers, "skip")
        assert is_sample_col >= 0
        assert hostname_col >= 0
        assert skip_col >= 0

        is_sample_values = [row[is_sample_col] for row in rows[1:]]
        assert "说明" in is_sample_values
        assert "样例" in is_sample_values

        data_rows = [r for r in rows[1:] if r[is_sample_col] is None or r[is_sample_col] == ""]
        assert len(data_rows) == 1
        assert data_rows[0][hostname_col] == "unknown.example.com"

    def test_skips_classified_ips(self, output_dir, rules_dir):
        data = {
            "1.1.1.1": {
                "classifier": {"category": "cloud_provider"},
                "rdns_ptr": {"has_ptr": True, "hostname": "ec2.amazonaws.com"},
            },
        }
        reader = _make_reader(data)

        count = export_unclassified_rdns(reader, output_dir, "test", rules_dir)
        assert count == 0
        assert not os.path.exists(os.path.join(output_dir, "test.unclassified_rdns.xlsx"))

    def test_skips_other_without_rdns(self, output_dir, rules_dir):
        data = {
            "1.1.1.1": {
                "classifier": {"category": "other"},
                "rdns_ptr": {"has_ptr": False},
            },
        }
        reader = _make_reader(data)

        count = export_unclassified_rdns(reader, output_dir, "test", rules_dir)
        assert count == 0

    def test_deduplicates_hostnames(self, output_dir, rules_dir):
        data = {
            "1.1.1.1": {
                "classifier": {"category": "other"},
                "rdns_ptr": {"has_ptr": True, "hostname": "same.example.com"},
            },
            "2.2.2.2": {
                "classifier": {"category": "other"},
                "rdns_ptr": {"has_ptr": True, "hostname": "same.example.com"},
            },
        }
        reader = _make_reader(data)

        count = export_unclassified_rdns(reader, output_dir, "test", rules_dir)
        assert count == 1

        sheets = _read_xlsx(os.path.join(output_dir, "test.unclassified_rdns.xlsx"))
        headers = sheets["未分类RDNS"][0]
        is_sample_col = _find_col(headers, "is_sample")
        hostname_col = _find_col(headers, "hostname")

        data_rows = [r for r in sheets["未分类RDNS"][1:] if r[is_sample_col] is None or r[is_sample_col] == ""]
        assert len(data_rows) == 1
        assert data_rows[0][hostname_col] == "same.example.com"

    def test_returns_zero_when_no_data(self, output_dir, rules_dir):
        reader = _make_reader({})
        count = export_unclassified_rdns(reader, output_dir, "test", rules_dir)
        assert count == 0


# ===== Import 行为：validate_row =====


class TestValidateRow:
    def test_valid_row_no_errors(self):
        row = {
            "is_sample": "",
            "category": "cloud_provider",
            "match_type": "suffix",
            "match_value": ".example.com",
        }
        errors = validate_row(row, {"cloud_provider"})
        assert errors == []

    def test_sample_rows_skipped(self):
        for marker in ("样例", "说明"):
            row = {"is_sample": marker, "category": "", "match_type": "", "match_value": ""}
            assert validate_row(row, set()) == []

    def test_skip_rows_skipped(self):
        row = {"is_sample": "", "skip": "跳过", "category": "", "match_type": "", "match_value": ""}
        assert validate_row(row, set()) == []

    def test_empty_category(self):
        row = {"is_sample": "", "category": "", "match_type": "suffix", "match_value": ".x.com"}
        errors = validate_row(row, set())
        assert any("category" in e for e in errors)

    def test_invalid_match_type(self):
        row = {
            "is_sample": "",
            "category": "cloud_provider",
            "match_type": "bad",
            "match_value": ".x.com",
        }
        errors = validate_row(row, {"cloud_provider"})
        assert any("match_type" in e for e in errors)

    def test_empty_match_value(self):
        row = {
            "is_sample": "",
            "category": "cloud_provider",
            "match_type": "suffix",
            "match_value": "",
        }
        errors = validate_row(row, {"cloud_provider"})
        assert any("match_value" in e for e in errors)

    def test_new_category_requires_extra_fields(self):
        row = {
            "is_sample": "",
            "category": "new_cat",
            "match_type": "suffix",
            "match_value": ".new.com",
            "new_label": "",
            "new_description": "",
            "new_need_deep_query": "",
        }
        errors = validate_row(row, {"cloud_provider"})
        assert any("new_label" in e for e in errors)
        assert any("new_description" in e for e in errors)
        assert any("new_need_deep_query" in e for e in errors)

    def test_new_category_valid_with_all_fields(self):
        row = {
            "is_sample": "",
            "category": "new_cat",
            "match_type": "suffix",
            "match_value": ".new.com",
            "new_label": "新分类",
            "new_description": "新分类描述",
            "new_need_deep_query": "是",
        }
        errors = validate_row(row, {"cloud_provider"})
        assert errors == []


# ===== Import 行为：merge_rules（纯函数） =====


class TestMergeRules:
    def test_appends_to_existing_category(self):
        existing = {"cloud_provider"}
        custom = OrderedDict()
        custom["cloud_provider"] = OrderedDict(
            [
                ("label", "云服务商"),
                ("description", "测试"),
                ("need_deep_query", True),
                (
                    "patterns",
                    [
                        {"field": "rdns_ptr.hostname", "match": ".amazonaws.com", "type": "suffix", "note": "AWS"},
                    ],
                ),
            ]
        )

        rows = [
            {
                "is_sample": "",
                "hostname": "rackspace.com",
                "field": "rdns_ptr.hostname",
                "category": "cloud_provider",
                "match_type": "suffix",
                "match_value": ".rackspace.com",
                "note": "Rackspace",
            },
        ]
        result, errors = merge_rules(rows, existing, custom)
        assert errors == []
        patterns = result["cloud_provider"]["patterns"]
        assert len(patterns) == 2
        assert patterns[1]["match"] == ".rackspace.com"

    def test_creates_new_category(self):
        existing = {"cloud_provider"}
        custom = OrderedDict()

        rows = [
            {
                "is_sample": "",
                "hostname": "dc.example.com",
                "field": "rdns_ptr.hostname",
                "category": "data_center",
                "match_type": "suffix",
                "match_value": ".dc.example.com",
                "note": "数据中心",
                "new_label": "数据中心",
                "new_description": "数据中心托管",
                "new_need_deep_query": "否",
            },
        ]
        result, errors = merge_rules(rows, existing, custom)
        assert errors == []
        assert "data_center" in result
        assert result["data_center"]["label"] == "数据中心"
        assert result["data_center"]["need_deep_query"] is False

    def test_skips_duplicate_match_value(self):
        existing = {"cloud_provider"}
        custom = OrderedDict()
        custom["cloud_provider"] = OrderedDict(
            [
                ("label", "云服务商"),
                ("description", "测试"),
                ("need_deep_query", True),
                (
                    "patterns",
                    [
                        {
                            "field": "rdns_ptr.hostname",
                            "match": ".rackspace.com",
                            "type": "suffix",
                            "note": "Rackspace",
                        },
                    ],
                ),
            ]
        )

        rows = [
            {
                "is_sample": "",
                "hostname": "rackspace.com",
                "field": "rdns_ptr.hostname",
                "category": "cloud_provider",
                "match_type": "suffix",
                "match_value": ".rackspace.com",
                "note": "重复",
            },
        ]
        result, errors = merge_rules(rows, existing, custom)
        assert errors == []
        assert len(result["cloud_provider"]["patterns"]) == 1

    def test_skips_sample_rows(self):
        existing = set()
        custom = OrderedDict()

        rows = [
            {
                "is_sample": "样例",
                "hostname": "test",
                "field": "rdns_ptr.hostname",
                "category": "cloud_provider",
                "match_type": "suffix",
                "match_value": ".test.com",
                "note": "",
            },
        ]
        result, errors = merge_rules(rows, existing, custom)
        assert errors == []
        assert "cloud_provider" not in result

    def test_skips_skip_rows(self):
        existing = set()
        custom = OrderedDict()

        rows = [
            {
                "is_sample": "",
                "skip": "跳过",
                "hostname": "test",
                "field": "rdns_ptr.hostname",
                "category": "cloud_provider",
                "match_type": "suffix",
                "match_value": ".test.com",
                "note": "重复主域名",
            },
        ]
        result, errors = merge_rules(rows, existing, custom)
        assert errors == []
        assert "cloud_provider" not in result

    def test_validation_errors_prevent_merge(self):
        existing = set()
        custom = OrderedDict()

        rows = [
            {
                "is_sample": "",
                "hostname": "test",
                "field": "rdns_ptr.hostname",
                "category": "",
                "match_type": "suffix",
                "match_value": ".test.com",
                "note": "",
            },
        ]
        result, errors = merge_rules(rows, existing, custom)
        assert len(errors) > 0
        assert "cloud_provider" not in result

    def test_mixed_valid_and_invalid_rows(self):
        existing = {"cloud_provider"}
        custom = OrderedDict()

        rows = [
            {
                "is_sample": "",
                "hostname": "good.com",
                "field": "rdns_ptr.hostname",
                "category": "cloud_provider",
                "match_type": "suffix",
                "match_value": ".good.com",
                "note": "有效",
            },
            {
                "is_sample": "",
                "hostname": "bad.com",
                "field": "rdns_ptr.hostname",
                "category": "",
                "match_type": "suffix",
                "match_value": ".bad.com",
                "note": "无效",
            },
        ]
        result, errors = merge_rules(rows, existing, custom)
        assert len(errors) == 1
        assert "cloud_provider" in result
        assert len(result["cloud_provider"]["patterns"]) == 1
        assert result["cloud_provider"]["patterns"][0]["match"] == ".good.com"
