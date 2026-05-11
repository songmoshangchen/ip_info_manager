import json
import logging
import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logger = logging.getLogger('ip_info_manager.scenarios.trace_ip')

HEADERS = [
    'IP', '国家', 'ASN/组织', '分类', '分类说明', '建议溯源路径',
    '域名数', '反查域名列表', '端口数', '开放端口列表', '标签',
]

SHEET_CONFIG = {
    1: {'title': 'P1 核心溯源', 'desc': '有反查域名 + 国内IP'},
    2: {'title': 'P2 重点溯源', 'desc': '有反查域名（国外）或无域名但有端口（国内）'},
    3: {'title': 'P3 辅助溯源', 'desc': '无域名但有端口（国外）或仅国内IP'},
    4: {'title': 'P4 暂缓', 'desc': '无域名、无端口、国外IP'},
}

LABEL_MAP = {
    'cloud_provider': '云服务商',
    'cdn': 'CDN/WAF节点',
    'crawler_scanner': '爬虫/扫描器',
    'residential': '家用宽带',
    'other': '其他（需确认）',
    'invalid_rdns': '无效RDNS',
    'excluded_domain': '排除域名',
}


def _extract_all_domains(info):
    domains = {}
    for src_name in ('aizhan', 'chinaz'):
        src = info.get(src_name, {})
        if not src.get('success'):
            continue
        for d in src.get('domains', []):
            domain = d.get('domain', '')
            if domain and domain not in domains:
                domains[domain] = True
    return list(domains.keys())


def _extract_fofa_ports(info):
    fofa = info.get('fofa_host', {})
    if fofa.get('error'):
        return []
    ports = []
    for p in fofa.get('ports', []):
        port_str = str(p.get('port', ''))
        products = ', '.join(
            pr.get('product', '') for pr in p.get('products', [])
        )
        if products:
            port_str = f"{port_str}({products})"
        ports.append(port_str)
    return ports


def _is_china_ip(info):
    ipinfo = info.get('ipinfo_api', {})
    return (ipinfo.get('country_code', '') == 'CN'
            or 'China' in ipinfo.get('country', ''))


def _has_domains(info):
    return len(_extract_all_domains(info)) > 0


def _has_ports(info):
    return len(_extract_fofa_ports(info)) > 0


def _trace_priority(ip, ip_data):
    is_cn = _is_china_ip(ip_data)
    has_dom = _has_domains(ip_data)
    has_pt = _has_ports(ip_data)
    if has_dom and is_cn:
        return 1
    if has_dom or (has_pt and is_cn):
        return 2
    if has_pt or is_cn:
        return 3
    return 4


def _sort_key(ip, ip_data):
    n_dom = len(_extract_all_domains(ip_data))
    n_pt = len(_extract_fofa_ports(ip_data))
    cat_weight = {'cloud_provider': 2, 'residential': 1, 'other': 0}
    cat = ip_data.get('trace_classify', {}).get('category', 'other')
    return (-n_dom, -n_pt, -cat_weight.get(cat, 0))


def _cat_display(info):
    classify = info.get('trace_classify', {})
    category = classify.get('category', '')
    if category == 'other':
        return LABEL_MAP.get('other', '其他（需确认）')
    label = LABEL_MAP.get(category, category)
    matched_by = classify.get('matched_by', [])
    if matched_by and matched_by[0].get('note'):
        note = matched_by[0]['note']
        return f'{label}（{note}）'
    return label


def _cat_note(info):
    classify = info.get('trace_classify', {})
    matched_by = classify.get('matched_by', [])
    if matched_by and matched_by[0].get('note'):
        return matched_by[0]['note']
    return ''


def _trace_action(info):
    has_dom = _has_domains(info)
    has_pt = _has_ports(info)
    is_cn = _is_china_ip(info)
    actions = []
    if has_dom:
        actions.append(
            'ICP备案/WHOIS查询域名注册信息' if is_cn
            else 'WHOIS查询域名注册信息')
    if has_pt:
        actions.append('排查端口服务泄露信息')
    if not actions:
        actions.append('公开信息检索IP历史行为')
    return '；'.join(actions)


def _format_domain_with_verify(domain, domain_verify):
    if not domain_verify or not domain_verify.get('results'):
        return domain
    for r in domain_verify['results']:
        if r['domain'] == domain:
            status = r.get('status', '')
            if status == 'matched':
                return f'{domain} ✅'
            elif status == 'changed':
                ips = ', '.join(r.get('resolved_ips', []))
                return f'{domain} 🔄→{ips}'
            elif status == 'unresolved':
                return f'{domain} ❌'
            elif status == 'timeout':
                return f'{domain} ⏱️'
            elif status == 'error':
                return f'{domain} ⚠️'
            break
    return domain


def _build_row(ip, info):
    country = info.get('ipinfo_api', {}).get('country', '')
    org = info.get('ipinfo_api', {}).get('as_name', '')
    domains = _extract_all_domains(info)
    ports = _extract_fofa_ports(info)
    tags_data = info.get('tags', [])
    tags_str = ', '.join(tags_data) if isinstance(tags_data, list) else str(tags_data)
    domain_verify = info.get('domain_verify')
    if domain_verify and domain_verify.get('results'):
        formatted_domains = [_format_domain_with_verify(d, domain_verify) for d in domains]
    else:
        formatted_domains = domains
    return [
        ip,
        country,
        org,
        _cat_display(info),
        _cat_note(info),
        _trace_action(info),
        str(len(domains)),
        '\n'.join(formatted_domains),
        str(len(ports)),
        '\n'.join(ports),
        tags_str,
    ]


def generate_trace_excel(output_dir, prefix):
    if not OPENPYXL_AVAILABLE:
        logger.warning('openpyxl 未安装，跳过 Excel 导出。安装命令：pip install openpyxl')
        return False

    json_path = os.path.join(output_dir, f'{prefix}.json')
    if not os.path.exists(json_path):
        logger.warning('找不到数据文件 %s，跳过 Excel 导出', json_path)
        return False

    with open(json_path, 'r', encoding='utf-8') as f:
        ip_data = json.load(f)

    deep_ips = []
    for ip in sorted(ip_data.keys()):
        if ip_data[ip].get('trace_classify', {}).get('need_deep_query'):
            deep_ips.append(ip)

    p_groups = {1: [], 2: [], 3: [], 4: []}
    for ip in deep_ips:
        lvl = _trace_priority(ip, ip_data[ip])
        p_groups[lvl].append(ip)

    for lvl in p_groups:
        p_groups[lvl].sort(key=lambda ip: _sort_key(ip, ip_data[ip]))

    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(
        start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_align = Alignment(
        horizontal='center', vertical='center', wrap_text=True)
    cell_align = Alignment(vertical='top', wrap_text=True)

    for lvl in [1, 2, 3, 4]:
        cfg = SHEET_CONFIG[lvl]
        ws = wb.create_sheet(title=cfg['title'])
        ws.append(HEADERS)

        for col_num in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        for ip in p_groups[lvl]:
            row = _build_row(ip, ip_data[ip])
            ws.append(row)
            row_num = ws.max_row
            for col_num in range(1, len(row) + 1):
                ws.cell(row=row_num, column=col_num).alignment = cell_align

        col_widths = {
            1: 18, 2: 10, 3: 25, 4: 22, 5: 20, 6: 30,
            7: 8, 8: 40, 9: 8, 10: 30, 11: 20,
        }
        for col_num, width in col_widths.items():
            ws.column_dimensions[chr(64 + col_num)].width = width

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = 'A2'

    output = os.path.join(output_dir, f'{prefix}.trace_report.xlsx')
    wb.save(output)
    logger.info('Excel 报告已生成：%s', output)
    return True
